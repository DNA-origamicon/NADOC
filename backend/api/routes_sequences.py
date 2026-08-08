"""
API layer — caDNAno-compatible sequence-file exports (extracted from crud.py).

This module hosts the routes that emit a per-strand *sequence* file for ordering
/ external tooling:

  - ``/design/export/sequence-csv`` — one row per strand (scaffold first),
    caDNAno-style CSV.
  - ``/design/export/sequence-xlsx`` — staple sequences as XLSX with 5′/3′
    overhang bases bolded and the on-screen strand color applied.

One reason to change: the file formats NADOC emits for strand-sequence ordering.
The *sequence-assignment* endpoints (assign scaffold / staple sequences) and the
*overhang random-sequence generation* routes are different concerns (they mutate
the design) and stay in crud.py.

The shared export resolver ``_design_for_export`` stays in crud.py (used across
crud.py + assembly.py + core) and is imported back here — same shared-kernel
convention as ``routes_export_structure.py`` / ``routes_camera_poses.py``.

URLs are unchanged from their previous home in crud.py. Mounting is done in
``backend/api/main.py`` via ``app.include_router(...)``.
"""

from __future__ import annotations

from fastapi import APIRouter, Body
from fastapi.responses import Response
from pydantic import BaseModel, Field

from backend.core.constants import STAPLE_PALETTE
from backend.core.models import StrandType
from backend.core.sequences import domain_bp_range, strand_nucleotide_count

# Shared export resolver used by many routes across crud.py + assembly.py + core;
# it stays in crud.py and is imported back here (same convention as
# routes_export_structure.py / routes_camera_poses.py).
from backend.api.crud import _design_for_export

router = APIRouter()


# ── caDNAno sequence export ────────────────────────────────────────────────────


@router.get("/design/export/sequence-csv")
def export_sequence_csv() -> Response:
    """Export strand sequences in caDNAno-compatible CSV format.

    Returns a CSV file matching caDNAno's staple export:
      Start, End, Sequence, Length, Color

    Scaffold and reference strands are excluded.
    """
    import csv
    import io
    import math

    from backend.core.cadnano import _assign_grid_coords, _HC_PERIOD, _SQ_PERIOD

    design = _design_for_export()

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["Start", "End", "Sequence", "Length", "Color"])

    helix_scaffold_dir = {h.id: None for h in design.helices}
    for strand in design.strands:
        if strand.strand_type != StrandType.SCAFFOLD:
            continue
        for domain in strand.domains:
            if domain.helix_id in helix_scaffold_dir:
                helix_scaffold_dir[domain.helix_id] = domain.direction

    rows, cols, export_dirs = _assign_grid_coords(
        design.helices, helix_scaffold_dir, design.lattice_type
    )
    sorted_helices = sorted(design.helices, key=lambda h: (rows[h.id], cols[h.id]))
    helix_num_map: dict[str, int] = {}
    fwd_i = rev_i = 0
    for h in sorted_helices:
        if export_dirs[h.id].value == "FORWARD":
            helix_num_map[h.id] = fwd_i * 2
            fwd_i += 1
        else:
            helix_num_map[h.id] = rev_i * 2 + 1
            rev_i += 1

    min_bp = 0
    seen_any = False
    for strand in design.strands:
        if strand.is_reference:
            continue
        for domain in strand.domains:
            if domain.helix_id not in helix_num_map:
                continue
            for bp in domain_bp_range(domain):
                min_bp = bp if not seen_any else min(min_bp, bp)
                seen_any = True
    for h in design.helices:
        for ls in h.loop_skips:
            min_bp = min(min_bp, ls.bp_index)

    period = _SQ_PERIOD if design.lattice_type.value == "SQUARE" else _HC_PERIOD
    offset = math.ceil((-min_bp) / period) * period if min_bp < 0 else 0

    def _endpoint(helix_id: str, bp: int) -> str:
        return f"{helix_num_map[helix_id]}[{bp + offset}]"

    def _sequence_for_export(strand, total_nt: int) -> str:
        if not strand.sequence:
            return "?" * total_nt
        seq = strand.sequence.replace(" ", "?")
        if len(seq) < total_nt:
            seq += "?" * (total_nt - len(seq))
        return seq

    strands_sorted = sorted(
        (
            s
            for s in design.strands
            if s.strand_type != StrandType.SCAFFOLD and not s.is_reference
        ),
        key=lambda s: (
            helix_num_map.get(s.domains[0].helix_id, 10**9) if s.domains else 10**9,
            s.domains[0].start_bp if s.domains else 10**9,
            s.id,
        ),
    )
    for strand in strands_sorted:
        if not strand.domains:
            continue
        first_d = strand.domains[0]
        last_d = strand.domains[-1]
        if (
            first_d.helix_id not in helix_num_map
            or last_d.helix_id not in helix_num_map
        ):
            continue
        # Loop/skip-adjusted nt count (NOT the raw bp-range span): deletions remove
        # bases, so the bp-range over-counts and pads the shorter real sequence with
        # spurious '?' — which crashes/skips CanDo. See strand_nucleotide_count.
        total_nt = strand_nucleotide_count(strand, design)
        seq = _sequence_for_export(strand, total_nt)
        color = (strand.color or "#f7931e").lower()
        writer.writerow(
            [
                _endpoint(first_d.helix_id, first_d.start_bp),
                _endpoint(last_d.helix_id, last_d.end_bp),
                seq,
                len(seq),
                color,
            ]
        )

    csv_bytes = output.getvalue().encode("utf-8")
    design_name = design.metadata.name or "design"
    filename = f"{design_name}_sequences.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class _SequenceXlsxRequest(BaseModel):
    strand_colors: dict[str, str] = Field(default_factory=dict)
    strand_order: list[str] = Field(default_factory=list)


@router.post("/design/export/sequence-xlsx")
def export_sequence_xlsx(
    req: _SequenceXlsxRequest | None = Body(default=None),
) -> Response:
    """Export staple sequences as XLSX with overhang regions bolded.

    Each Sequence cell is rich-text: 5′/3′ overhang bases are bold, the body
    is plain.  All three segments share the effective strand color (provided
    via ``strand_colors`` to match the on-screen Sequence panel; falls back
    to ``strand.color`` or the staple palette).  Strand order can be supplied
    via ``strand_order`` to match the panel sort.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Font, Alignment

    design = _design_for_export()
    color_overrides = (req.strand_colors if req else {}) or {}
    order = (req.strand_order if req else []) or []

    def _hex_to_argb(hexstr: str) -> str:
        h = (hexstr or "").lstrip("#")
        if len(h) != 6:
            return "FF000000"
        return "FF" + h.upper()

    helix_label_by_id: dict[str, str] = {}
    for idx, h in enumerate(design.helices):
        helix_label_by_id[h.id] = h.label if h.label else str(idx)

    wb = Workbook()
    ws = wb.active
    ws.title = "Staples"
    headers = ["#", "Sequence", "Length", "Color", "Start", "End", "Notes"]
    ws.append(headers)
    hf = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = hf

    # Palette fallback must agree with the Sequence panel (ui/spreadsheet.js
    # `paletteColor`), which indexes the CANONICAL STAPLE_PALETTE by the
    # strand's position in `design.strands` — not by the sorted row number.
    # This branch only fires for headless/API exports; the UI supplies
    # `strand_colors` for every strand.
    strand_pos = {s.id: i for i, s in enumerate(design.strands)}

    staples = [s for s in design.strands if s.strand_type != StrandType.SCAFFOLD]
    if order:
        pos = {sid: i for i, sid in enumerate(order)}
        staples.sort(key=lambda s: pos.get(s.id, len(order)))

    for row_idx, strand in enumerate(staples, start=1):
        if not strand.domains:
            continue
        d0, dn = strand.domains[0], strand.domains[-1]
        ovhg5_len = (abs(d0.end_bp - d0.start_bp) + 1) if d0.overhang_id else 0
        ovhg3_len = (abs(dn.end_bp - dn.start_bp) + 1) if dn.overhang_id else 0
        total_len = strand_nucleotide_count(
            strand, design
        )  # loop/skip-adjusted (see CSV export)

        color_hex = (
            color_overrides.get(strand.id)
            or strand.color
            or STAPLE_PALETTE[strand_pos.get(strand.id, 0) % len(STAPLE_PALETTE)]
        )
        argb = _hex_to_argb(color_hex)

        ws.cell(row=row_idx + 1, column=1, value=row_idx)

        seq = strand.sequence or ""
        if seq:
            ov5 = seq[:ovhg5_len] if ovhg5_len else ""
            ov3 = seq[len(seq) - ovhg3_len :] if ovhg3_len else ""
            body_seq = (
                seq[ovhg5_len : len(seq) - ovhg3_len]
                if (ovhg5_len + ovhg3_len) < len(seq)
                else ""
            )
            bold_font = InlineFont(rFont="Courier New", b=True, color=argb)
            plain_font = InlineFont(rFont="Courier New", color=argb)
            blocks: list[TextBlock] = []
            if ov5:
                blocks.append(TextBlock(bold_font, ov5))
            if body_seq:
                blocks.append(TextBlock(plain_font, body_seq))
            if ov3:
                blocks.append(TextBlock(bold_font, ov3))
            ws.cell(
                row=row_idx + 1, column=2, value=CellRichText(blocks) if blocks else ""
            )
        else:
            # No sequence assigned — show N×length unbolded so the column isn't empty
            ws.cell(row=row_idx + 1, column=2, value=f"N×{total_len}")

        ws.cell(row=row_idx + 1, column=3, value=total_len)
        ws.cell(row=row_idx + 1, column=4, value=color_hex)
        ws.cell(
            row=row_idx + 1,
            column=5,
            value=f"{helix_label_by_id.get(d0.helix_id, d0.helix_id)}[{d0.start_bp}]",
        )
        ws.cell(
            row=row_idx + 1,
            column=6,
            value=f"{helix_label_by_id.get(dn.helix_id, dn.helix_id)}[{dn.end_bp}]",
        )
        ws.cell(row=row_idx + 1, column=7, value=strand.notes or "")

    for col_letter, w in (
        ("A", 6),
        ("B", 80),
        ("C", 8),
        ("D", 10),
        ("E", 12),
        ("F", 12),
        ("G", 30),
    ):
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "A2"
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="left")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    design_name = design.metadata.name or "design"
    filename = f"{design_name}_sequences.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
