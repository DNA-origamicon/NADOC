from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api import state as design_state
from backend.api.main import app
from backend.api.routes import _demo_design
from backend.core.models import PlateLayout, TubeAssignment, WellAssignment

client = TestClient(app)


@pytest.fixture(autouse=True)
def _reset_design():
    design_state.set_design(_demo_design())
    yield
    design_state.set_design(_demo_design())


def test_idt_order_xlsx_uses_saved_wells_and_separate_sheets():
    design = _demo_design()
    staples = [s for s in design.strands if s.strand_type.value == "staple"]
    second = staples[0].model_copy(deep=True, update={"id": "staple_1"})
    design.strands.append(second)
    staples.append(second)
    staples[0].sequence = "ACGT"
    staples[1].sequence = "TGCA"
    design.plate_layout = PlateLayout(
        plate_count=2,
        wells=[
            WellAssignment(strand_id=staples[0].id, plate=0, row=1, col=2),
            WellAssignment(strand_id=staples[1].id, plate=1, row=7, col=11),
        ],
        tubes=[],
    )
    design_state.set_design(design)

    response = client.post(
        "/api/design/export/idt-order-xlsx",
        json={"strand_names": {staples[0].id: "Body_1", staples[1].id: "Body_2"}},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Plate 1", "Plate 2"]
    assert list(wb["Plate 1"].values) == [("Well", "Name", "Sequence"), ("B3", "Body_1", "ACGT")]
    assert list(wb["Plate 2"].values) == [("Well", "Name", "Sequence"), ("H12", "Body_2", "TGCA")]


def test_idt_order_xlsx_puts_tubes_on_a_tubes_sheet():
    design = _demo_design()
    staple = next(s for s in design.strands if s.strand_type.value == "staple")
    staple.sequence = "AAAA"
    design.plate_layout = PlateLayout(
        wells=[], tubes=[TubeAssignment(strand_id=staple.id, reason="long")]
    )
    design_state.set_design(design)

    response = client.post(
        "/api/design/export/idt-order-xlsx",
        json={"strand_names": {staple.id: "Long_1"}},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Plate 1", "Tubes"]
    assert list(wb["Tubes"].values) == [("Well", "Name", "Sequence"), (None, "Long_1", "AAAA")]
