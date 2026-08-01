"""
Regression pin for the XLSX oligo-order sheet's staple-colour fallback (TD-02).

`POST /api/design/export/sequence-xlsx` hard-coded its own 12-colour `_PALETTE` —
the *retired* editor syntax theme (`#e06c75 #98c379 #d19a66 #61afef …`) that
`ui/spreadsheet.js` was purged of on 2026-07-30. The frontend supplies
`strand_colors` for every strand, so the UI export looked correct and the bug
hid; but any headless / API-driven export fell through to that palette and
produced an order sheet whose colours matched nothing in the app.

It also indexed by *sorted row position* rather than the strand's position in
`design.strands`, so even the hue-to-strand mapping disagreed with the panel.

These assertions fail against the old code: it emitted '#e06c75' for row 1.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api import state as design_state
from backend.api.main import app
from backend.core.constants import STAPLE_PALETTE
from backend.core.models import StrandType
from tests.conftest import make_6hb_design

client = TestClient(app)

# The palette the export used to carry. Nothing in the app uses these any more.
RETIRED_SYNTAX_THEME = {
    "#e06c75", "#98c379", "#d19a66", "#61afef",
    "#c678dd", "#56b6c2", "#e5c07b", "#abb2bf",
    "#be5046", "#7dab6e", "#b07e45", "#4e8cc4",
}


@pytest.fixture
def staples_without_colour():
    """A 6HB whose staples carry no explicit colour — the Full-Autostaple case."""
    design = make_6hb_design()
    for strand in design.strands:
        strand.color = None
    design_state.set_design(design)
    staples = [s for s in design.strands if s.strand_type != StrandType.SCAFFOLD]
    if not staples:
        pytest.skip("fixture has no staple strands to colour")
    yield design
    design_state.set_design(make_6hb_design())


def _colour_column(response) -> list[str]:
    """Column D ('Color') of every data row, as written by the export."""
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    return [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]


def test_xlsx_fallback_uses_the_canonical_staple_palette(staples_without_colour):
    r = client.post("/api/design/export/sequence-xlsx", json={"strand_colors": {}, "strand_order": []})
    assert r.status_code == 200

    colours = _colour_column(r)
    assert colours, "export produced no data rows"
    for hex_str in colours:
        assert hex_str in STAPLE_PALETTE, f"{hex_str} is not a canonical palette colour"
        assert hex_str not in RETIRED_SYNTAX_THEME


def test_xlsx_fallback_indexes_by_design_strands_position(staples_without_colour):
    """The slot must come from the strand's index in `design.strands` — the same
    scheme `ui/spreadsheet.js` uses — not from the export's own row counter."""
    design = staples_without_colour
    expected = {
        s.id: STAPLE_PALETTE[i % len(STAPLE_PALETTE)]
        for i, s in enumerate(design.strands)
        if s.strand_type != StrandType.SCAFFOLD and s.domains
    }
    # Force panel order == design order so row N ↔ the Nth staple.
    order = [sid for sid in expected]
    r = client.post(
        "/api/design/export/sequence-xlsx",
        json={"strand_colors": {}, "strand_order": order},
    )
    assert r.status_code == 200
    assert _colour_column(r) == [expected[sid] for sid in order]


def test_explicit_strand_colours_still_win(staples_without_colour):
    design = staples_without_colour
    first = next(s for s in design.strands if s.strand_type != StrandType.SCAFFOLD and s.domains)
    r = client.post(
        "/api/design/export/sequence-xlsx",
        json={"strand_colors": {first.id: "#123456"}, "strand_order": [first.id]},
    )
    assert r.status_code == 200
    assert _colour_column(r)[0] == "#123456"
